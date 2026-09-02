from __future__ import annotations

import unittest
from pathlib import Path

from openpyxl import load_workbook

from common.no_requirements import (
    ADJUDICATION_BY_LINE,
    ADJUDICATION_GLOBAL,
    ADJUDICATION_UNKNOWN,
    NO_REQUIREMENTS_MIXED,
    NO_REQUIREMENTS_ONLY,
    NO_REQUIREMENTS_SCOPE_COLUMN,
    adjudication_type_from_cells,
    adjudication_type_from_detail_text,
    classify_no_requirements_scope,
    evaluate_no_requirements,
    ficha_codes_from_label,
    normalize_adjudication_type,
    resolve_adjudication_type,
    scope_column_values,
)


class NoRequirementsClassificationTests(unittest.TestCase):
    def setUp(self) -> None:
        self.no_req = {"100", "200", "51119"}
        self.ct = {"300"}
        self.rs = {"400"}

    def classify(self, fichas) -> str:
        return classify_no_requirements_scope(
            fichas,
            no_requirements=self.no_req,
            requires_ct=self.ct,
            requires_rs=self.rs,
        )

    def test_only_no_requirements(self) -> None:
        self.assertEqual(self.classify(["100", "*200"]), NO_REQUIREMENTS_ONLY)

    def test_mixed_with_ct_rs_or_unknown(self) -> None:
        self.assertEqual(self.classify(["100", "300"]), NO_REQUIREMENTS_MIXED)
        self.assertEqual(self.classify(["100", "400"]), NO_REQUIREMENTS_MIXED)
        self.assertEqual(self.classify(["100", "999"]), NO_REQUIREMENTS_MIXED)

    def test_without_eligible_ficha_is_not_labeled(self) -> None:
        self.assertEqual(self.classify(["300"]), "")

    def test_contradictory_catalog_is_conservative(self) -> None:
        self.assertEqual(
            classify_no_requirements_scope(
                ["51119"],
                no_requirements=self.no_req,
                requires_ct={"51119"},
                requires_rs=set(),
            ),
            NO_REQUIREMENTS_MIXED,
        )

    def test_historical_sheet_values(self) -> None:
        rows = [
            ["ficha_detectada", NO_REQUIREMENTS_SCOPE_COLUMN],
            ["100, *200", ""],
            ["100, 300", NO_REQUIREMENTS_ONLY],
        ]
        values, changes = scope_column_values(
            rows,
            no_requirements=self.no_req,
            requires_ct=self.ct,
            requires_rs=self.rs,
        )
        self.assertEqual(values, [[NO_REQUIREMENTS_ONLY], [NO_REQUIREMENTS_MIXED]])
        self.assertEqual(changes, 2)

    def test_ficha_label_parser_preserves_order(self) -> None:
        self.assertEqual(ficha_codes_from_label("*100, * 200; 100"), ["100", "200"])

    def decision(self, fichas, mode):
        return evaluate_no_requirements(
            fichas,
            no_requirements=self.no_req,
            requires_ct=self.ct,
            requires_rs=self.rs,
            adjudication_type=mode,
        )

    def test_pure_act_is_eligible_with_any_adjudication(self) -> None:
        decision = self.decision(["100", "200"], "Global")
        self.assertTrue(decision.eligible)
        self.assertEqual(decision.scope, NO_REQUIREMENTS_ONLY)
        self.assertEqual(decision.no_requirements_fichas, ("100", "200"))

    def test_mixed_act_requires_line_adjudication(self) -> None:
        by_line = self.decision(["100", "300"], "Renglón")
        global_act = self.decision(["100", "300"], "Global")
        unknown = self.decision(["100", "300"], "")

        self.assertTrue(by_line.eligible)
        self.assertFalse(global_act.eligible)
        self.assertFalse(unknown.eligible)
        self.assertEqual(by_line.requirements_label, "300 (CT)")

    def test_rs_and_combined_requirements_are_labeled(self) -> None:
        decision = evaluate_no_requirements(
            ["100", "300", "400", "500"],
            no_requirements=self.no_req,
            requires_ct={"300", "500"},
            requires_rs={"400", "500"},
            adjudication_type="Por renglones",
        )
        self.assertEqual(decision.requirements_fichas, ("300", "400", "500"))
        self.assertEqual(decision.requirements_label, "300 (CT), 400 (RS), 500 (CT/RS)")
        self.assertTrue(decision.eligible)

    def test_unknown_ficha_is_mixed_and_fails_closed_without_line_mode(self) -> None:
        decision = self.decision(["100", "999"], "Global")
        self.assertEqual(decision.scope, NO_REQUIREMENTS_MIXED)
        self.assertEqual(decision.unclassified_fichas, ("999",))
        self.assertFalse(decision.eligible)

    def test_adjudication_normalization(self) -> None:
        self.assertEqual(normalize_adjudication_type("Renglón"), ADJUDICATION_BY_LINE)
        self.assertEqual(normalize_adjudication_type("Adjudicación parcial"), ADJUDICATION_BY_LINE)
        self.assertEqual(normalize_adjudication_type("por ítem"), ADJUDICATION_BY_LINE)
        self.assertEqual(normalize_adjudication_type("GLOBAL"), ADJUDICATION_GLOBAL)
        self.assertEqual(normalize_adjudication_type("Por lote"), ADJUDICATION_UNKNOWN)
        self.assertEqual(
            adjudication_type_from_cells(["acto", "fecha", "Global"]),
            ADJUDICATION_GLOBAL,
        )

    def test_detail_fallback_is_anchored_to_the_field_label(self) -> None:
        detail = """Proveedor Global Corp\nModalidad de adjudicación\nRenglón\nMonto"""
        self.assertEqual(
            adjudication_type_from_detail_text(detail),
            ADJUDICATION_BY_LINE,
        )
        self.assertEqual(
            adjudication_type_from_detail_text("Proveedor Global Corp\nMonto"),
            ADJUDICATION_UNKNOWN,
        )

    def test_listing_value_has_precedence_over_detail_fallback(self) -> None:
        self.assertEqual(
            resolve_adjudication_type("Global", "Modalidad de adjudicación\nRenglón"),
            ADJUDICATION_GLOBAL,
        )


class NoRequirementsCatalogTests(unittest.TestCase):
    TARGETS = {"51119", "51138", "51333", "51334", "51460"}
    ROOT = Path(__file__).resolve().parents[1]

    @staticmethod
    def _rows(path: Path) -> list[list[str]]:
        workbook = load_workbook(path, read_only=True, data_only=True)
        worksheet = workbook.active
        return [
            [str(value).strip() if value is not None else "" for value in row]
            for row in worksheet.iter_rows(values_only=True)
        ]

    def test_corrected_fichas_are_only_in_no_requirements_catalog(self) -> None:
        no_req = self._rows(self.ROOT / "data/fichas/fichas_sin_requisitos.xlsx")
        ct = self._rows(self.ROOT / "data/clrir/fichas_con_CT_sin_RS.xlsx")
        rs = self._rows(self.ROOT / "data/clrir/fichas_con_RS.xlsx")
        con_req = self._rows(self.ROOT / "data/fichas/fichas_con_requisitos.xlsx")

        no_req_codes = [row[0] for row in no_req if row and row[0] in self.TARGETS]
        self.assertEqual(set(no_req_codes), self.TARGETS)
        self.assertEqual(len(no_req_codes), len(self.TARGETS))
        self.assertFalse(self.TARGETS.intersection(row[0] for row in ct if row))
        self.assertFalse(self.TARGETS.intersection(row[0] for row in rs if row))
        self.assertFalse(self.TARGETS.intersection(row[0] for row in con_req if row))

    def test_corrected_fichas_have_one_no_no_row_in_master_catalog(self) -> None:
        master = self._rows(self.ROOT / "data/fichas/todas_las_fichas.xlsx")
        selected = [row for row in master if row and row[0] in self.TARGETS]
        self.assertEqual(len(selected), len(self.TARGETS))
        self.assertEqual({row[0] for row in selected}, self.TARGETS)
        self.assertTrue(all(row[1:3] == ["NO", "NO"] for row in selected))


if __name__ == "__main__":
    unittest.main()
